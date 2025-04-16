from fastapi import FastAPI, HTTPException, Depends, status, Form, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, Dict, List
from datetime import datetime, date, timedelta
import sqlite3
from jose import jwt
import bcrypt
from pathlib import Path
import os
from fastapi import Request
import pandas as pd
import io
import tempfile
import secrets
import xlsxwriter
import logging
from openpyxl import Workbook
from io import BytesIO
from urllib.parse import quote
import re

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = FastAPI()

# 自动生成 SECRET_KEY
SECRET_KEY = secrets.token_hex(32)
logger.info(f"Generated new SECRET_KEY: {SECRET_KEY}")

# 配置CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有来源，包括局域网
    allow_credentials=True,
    allow_methods=["*"],  # 允许所有方法
    allow_headers=["*"],  # 允许所有头部
)

# 配置模板目录
STATIC_DIR = str(Path(__file__).resolve().parent.parent / "frontend" / "static")
TEMPLATES_DIR = str(Path(__file__).resolve().parent.parent / "frontend" / "templates")

# 挂载静态文件
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# 配置模板
templates = Jinja2Templates(directory=TEMPLATES_DIR)

# 数据库配置
BASE_DIR = Path(__file__).resolve().parent
DATABASE = str(BASE_DIR / "duty_system.db")

# 数据库连接参数
DB_PARAMS = {
    "check_same_thread": False,
    "timeout": 30,
    "isolation_level": None,  # 自动提交模式
    "cached_statements": 100,  # 缓存SQL语句
    "uri": True  # 启用URI模式
}

# 数据库连接
def get_db():
    try:
        conn = sqlite3.connect(f"file:{DATABASE}?cache=shared", **DB_PARAMS)
        conn.row_factory = sqlite3.Row
        yield conn
    except sqlite3.Error as e:
        logger.error(f"Database connection error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Database connection error"
        )
    finally:
        if 'conn' in locals():
            conn.close()

# 初始化数据库
def init_db():
    try:
        with sqlite3.connect(f"file:{DATABASE}?cache=shared", **DB_PARAMS) as conn:
            cursor = conn.cursor()
            
            # 创建用户表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    department TEXT,
                    is_admin INTEGER DEFAULT 0,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 创建用户表索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_users_username ON users(username)")
            
            # 检查管理员账号是否存在
            cursor.execute("SELECT username FROM users WHERE username = 'admin'")
            if not cursor.fetchone():
                # 创建管理员账号
                hashed_password = bcrypt.hashpw("admin123".encode(), bcrypt.gensalt())
                cursor.execute("""
                    INSERT INTO users (username, password, is_admin)
                    VALUES (?, ?, 1)
                """, ("admin", hashed_password))
            
            # 创建值班信息表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS duty_info (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    department TEXT NOT NULL,
                    leader_name TEXT,
                    leader_phone TEXT,
                    manager_name TEXT,
                    manager_phone TEXT,
                    member_name TEXT,
                    member_phone TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 创建值班信息表索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_duty_info_date ON duty_info(date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_duty_info_department ON duty_info(department)")
            
            # 创建运营中心固定值班人员表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operation_center_fixed_duty (
                    shift INTEGER PRIMARY KEY,
                    leader_name TEXT,
                    leader_phone TEXT,
                    manager_name TEXT,
                    manager_phone TEXT,
                    member_name TEXT,
                    member_phone TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 创建运营中心值班人员表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operation_center_duty (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    shift INTEGER NOT NULL,
                    backup_name TEXT,
                    backup_phone TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(date, shift)
                )
            """)
            
            # 创建运营中心值班人员表索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_operation_center_duty_date ON operation_center_duty(date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_operation_center_duty_shift ON operation_center_duty(shift)")
            
            conn.commit()
            logger.info("Database initialized successfully")
    except sqlite3.Error as e:
        logger.error(f"Database initialization error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Database initialization error"
        )

# 获取东八区时间
def get_east8_time():
    """获取东八区时间"""
    utc_now = datetime.utcnow()
    east8_time = utc_now + timedelta(hours=8)
    return east8_time.strftime("%Y-%m-%d %H:%M:%S")

# 工具函数
def verify_password(plain_password, hashed_password):
    try:
        if isinstance(hashed_password, str):
            hashed_password = hashed_password.encode()
        return bcrypt.checkpw(plain_password.encode(), hashed_password)
    except (ValueError, TypeError) as e:
        logger.error(f"Password verification error: Invalid password format - {str(e)}")
        return False
    except Exception as e:
        logger.error(f"Password verification error: Unexpected error - {str(e)}")
        return False

def get_user(username: str):
    try:
        with sqlite3.connect(f"file:{DATABASE}?cache=shared", **DB_PARAMS) as conn:
            c = conn.cursor()
            c.execute("SELECT username, password, department, is_admin FROM users WHERE username = ?", (username,))
            user = c.fetchone()
            if user:
                return UserInDB(
                    username=user[0],
                    password=user[1],
                    department=user[2],
                    is_admin=bool(user[3])
                )
            return None
    except sqlite3.Error as e:
        logger.error(f"Database error in get_user: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error in get_user: {str(e)}")
        return None

# 模型定义
class User(BaseModel):
    username: str
    department: Optional[str] = None
    is_admin: bool = False

class UserInDB(User):
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class DutyInfo(BaseModel):
    department: str
    date: str
    leader_name: Optional[str] = None
    leader_phone: Optional[str] = None
    manager_name: Optional[str] = None
    manager_phone: Optional[str] = None
    member_name: Optional[str] = None
    member_phone: Optional[str] = None

class OperationCenterDuty(BaseModel):
    date: str
    shift: int
    backup_name: Optional[str] = None
    backup_phone: Optional[str] = None

class OperationCenterFixedDuty(BaseModel):
    shift: int
    leader_name: str
    leader_phone: str
    manager_name: str
    manager_phone: str
    member_name: str
    member_phone: str

class OperationCenterDutyHistory(BaseModel):
    start_date: str
    end_date: str
    shift: Optional[int] = None

class OperationCenterDutyBatch(BaseModel):
    shifts: List[int]
    leader_name: str
    leader_phone: str
    manager_name: str
    manager_phone: str
    member_name: str
    member_phone: str

# 安全配置
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def authenticate_user(username: str, password: str):
    try:
        user = get_user(username)
        if not user:
            logger.warning(f"Authentication failed: User {username} not found")
            return False
        if not verify_password(password, user.password):
            logger.warning(f"Authentication failed: Invalid password for user {username}")
            return False
        logger.info(f"User {username} authenticated successfully")
        return user
    except Exception as e:
        logger.error(f"Authentication error: {str(e)}")
        return False

def create_access_token(data: dict):
    to_encode = data.copy()
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), db: sqlite3.Connection = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="无效的认证凭据",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except jwt.JWTError:
        raise credentials_exception
    
    try:
        cursor = db.cursor()
        cursor.execute("SELECT username, department, is_admin FROM users WHERE username = ?", (username,))
        user = cursor.fetchone()
        if user is None:
            raise credentials_exception
        return User(username=user[0], department=user[1], is_admin=bool(user[2]))
    except sqlite3.Error as e:
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")

# API路由
@app.post("/token")
async def login(
    form_data: OAuth2PasswordRequestForm = Depends()
):
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, department, is_admin FROM users WHERE username = ?", (form_data.username,))
        user = cursor.fetchone()
        
        if not user:
            raise HTTPException(status_code=401, detail="用户名或密码错误")
        
        # 验证密码
        stored_password = user[1]  # password在第二列
        
        # 确保 stored_password 是字节类型
        if isinstance(stored_password, str):
            stored_password = stored_password.encode()
        
        # 密码验证
        password_valid = False
        try:
            password_valid = bcrypt.checkpw(form_data.password.encode(), stored_password)
        except Exception as e:
            print(f"密码验证错误: {str(e)}")
            # 如果验证出错，尝试重新初始化管理员密码
            if form_data.username == "admin":
                # 重置管理员密码
                hashed_password = bcrypt.hashpw("admin123".encode(), bcrypt.gensalt())
                cursor.execute("UPDATE users SET password = ? WHERE username = 'admin'", (hashed_password,))
                conn.commit()
                # 重新验证
                password_valid = bcrypt.checkpw(form_data.password.encode(), hashed_password)
        
        if not password_valid:
            raise HTTPException(status_code=401, detail="用户名或密码错误")
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = jwt.encode(
            {
                "sub": user[0],  # username在第一列
                "is_admin": bool(user[3]),  # is_admin在第四列
                "department": user[2],  # department在第三列
                "exp": datetime.utcnow() + access_token_expires
            },
            SECRET_KEY,
            algorithm=ALGORITHM
        )
        
        return {"access_token": access_token, "token_type": "bearer", "is_admin": bool(user[3])}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Login error: {str(e)}")  # 添加错误日志
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/duty-info")
async def get_duty_info(date: date, db: sqlite3.Connection = Depends(get_db)):
    try:
        c = db.cursor()
        
        # 获取所有部门（包括历史记录中的部门）
        c.execute("""
            SELECT DISTINCT department 
            FROM duty_info 
            WHERE date <= ?
            ORDER BY department
        """, (date.isoformat(),))
        
        all_departments = [row[0] for row in c.fetchall()]
        
        # 获取指定日期的值班信息
        c.execute("""
            SELECT department, leader_name, leader_phone, manager_name, manager_phone, 
                   member_name, member_phone
            FROM duty_info
            WHERE date = ?
        """, (date.isoformat(),))
        
        duty_info = {}
        # 首先为所有部门创建空记录
        for dept in all_departments:
            duty_info[dept] = {
                "leader_name": "",
                "leader_phone": "",
                "manager_name": "",
                "manager_phone": "",
                "member_name": "",
                "member_phone": ""
            }
        
        # 然后填充有值班信息的部门
        for row in c.fetchall():
            duty_info[row[0]] = {
                "leader_name": row[1],
                "leader_phone": row[2],
                "manager_name": row[3],
                "manager_phone": row[4],
                "member_name": row[5],
                "member_phone": row[6]
            }
            
        return duty_info
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取值班信息失败: {str(e)}"
        )

@app.post("/api/duty-info")
async def create_duty_info(
    duty_info: DutyInfo,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    if not current_user.is_admin and current_user.department != duty_info.department:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="没有权限修改该部门的值班信息"
        )
    
    try:
        print(f"更新值班信息: 部门={duty_info.department}, 日期={duty_info.date}")
        
        # 首先检查记录是否存在
        c = db.cursor()
        c.execute("""
            SELECT COUNT(*) FROM duty_info 
            WHERE date = ? AND department = ?
        """, (duty_info.date, duty_info.department))
        
        count = c.fetchone()[0]
        current_time = datetime.now().isoformat()
        
        if count > 0:
            # 更新现有记录
            c.execute("""
                UPDATE duty_info 
                SET leader_name = ?, leader_phone = ?,
                    manager_name = ?, manager_phone = ?,
                    member_name = ?, member_phone = ?,
                    updated_at = ?
                WHERE date = ? AND department = ?
            """, (
                duty_info.leader_name,
                duty_info.leader_phone,
                duty_info.manager_name,
                duty_info.manager_phone,
                duty_info.member_name,
                duty_info.member_phone,
                current_time,
                duty_info.date,
                duty_info.department
            ))
        else:
            # 插入新记录
            c.execute("""
                INSERT INTO duty_info 
                (date, department, leader_name, leader_phone, 
                 manager_name, manager_phone, member_name, member_phone,
                 created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                duty_info.date,
                duty_info.department,
                duty_info.leader_name,
                duty_info.leader_phone,
                duty_info.manager_name,
                duty_info.manager_phone,
                duty_info.member_name,
                duty_info.member_phone,
                current_time,
                current_time
            ))
        
        db.commit()
        return {"message": "值班信息已更新"}
    except sqlite3.IntegrityError as e:
        print(f"完整性错误: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"更新值班信息失败: {str(e)}"
        )
    except sqlite3.OperationalError as e:
        print(f"操作错误: {str(e)}")
        if "database is locked" in str(e):
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="数据库正忙，请稍后重试"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新值班信息失败: {str(e)}"
        )
    except Exception as e:
        print(f"其他错误: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新值班信息失败: {str(e)}"
        )

@app.delete("/api/duty-info/{department}")
async def delete_duty_info(
    department: str,
    date: date,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    if not current_user.is_admin and current_user.department != department:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="没有权限删除该部门的值班信息"
        )
    
    if date > date.today():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="不能删除未来日期的值班信息"
        )
    
    c = db.cursor()
    c.execute("""
        DELETE FROM duty_info
        WHERE department = ? AND date = ?
    """, (department, date.isoformat()))
    db.commit()
    return {"message": "值班信息已删除"}

@app.get("/api/users/me", response_model=User)
async def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

# 部门管理API
@app.get("/api/departments")
async def get_departments(db: sqlite3.Connection = Depends(get_db)):
    """获取所有部门列表"""
    try:
        cursor = db.cursor()
        cursor.execute("""
            SELECT DISTINCT department 
            FROM duty_info 
            WHERE date >= date('now')
            ORDER BY department
        """)
        departments = [row[0] for row in cursor.fetchall()]
        return departments
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/departments")
async def create_department(department: dict, current_user: User = Depends(get_current_user)):
    """创建新部门"""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="只有管理员可以创建部门")
    
    conn = None
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 检查部门是否已存在
        cursor.execute("""
            SELECT COUNT(*) FROM duty_info 
            WHERE department = ? AND date >= ?
        """, (department["name"], department["start_date"]))
        
        if cursor.fetchone()[0] > 0:
            raise HTTPException(status_code=400, detail="该部门在指定日期后已存在")
        
        # 获取所有未来的日期
        cursor.execute("""
            SELECT DISTINCT date FROM duty_info 
            WHERE date >= ? 
            ORDER BY date
        """, (department["start_date"],))
        
        future_dates = [row[0] for row in cursor.fetchall()]
        
        # 如果没有未来的日期，至少创建一条记录
        if not future_dates:
            cursor.execute("""
                INSERT INTO duty_info 
                (date, department, leader_name, leader_phone, manager_name, manager_phone, member_name, member_phone)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                department["start_date"],
                department["name"],
                None, None, None, None, None, None
            ))
        else:
            # 为每个未来日期创建记录
            for date in future_dates:
                cursor.execute("""
                    INSERT INTO duty_info 
                    (date, department, leader_name, leader_phone, manager_name, manager_phone, member_name, member_phone)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    date,
                    department["name"],
                    None, None, None, None, None, None
                ))
        
        conn.commit()
        return {"message": "部门创建成功"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            conn.close()

@app.put("/api/departments/{old_name}")
async def update_department(old_name: str, department: dict, current_user: User = Depends(get_current_user)):
    """更新部门名称（只更新指定日期及以后的记录）"""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="只有管理员可以修改部门")
    
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 更新值班信息表中的部门名称
        cursor.execute("""
            UPDATE duty_info 
            SET department = ? 
            WHERE department = ? AND date >= ?
        """, (department["name"], old_name, department.get("start_date", datetime.now().date().isoformat())))
        
        # 更新用户表中的部门名称
        cursor.execute("""
            UPDATE users 
            SET department = ? 
            WHERE department = ?
        """, (department["name"], old_name))
        
        conn.commit()
        return {"message": "部门更新成功"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            conn.close()

@app.delete("/api/departments/{name}")
async def delete_department(name: str, date: str, current_user: User = Depends(get_current_user)):
    """删除部门（只删除指定日期及以后的记录）"""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="只有管理员可以删除部门")
    
    conn = None
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 只删除指定日期及以后的记录
        cursor.execute("""
            DELETE FROM duty_info 
            WHERE department = ? AND date >= ?
        """, (name, date))
        
        conn.commit()
        return {"message": "部门删除成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            conn.close()

@app.get("/api/duty-info/{department}")
async def get_department_duty_info(
    department: str, 
    date: str,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """获取指定部门在指定日期的值班信息"""
    try:
        print(f"获取值班信息: 部门={department}, 日期={date}")
        
        # 检查权限
        if not current_user.is_admin and current_user.department != department:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="没有权限访问该部门的值班信息"
            )
        
        c = db.cursor()
        c.execute("""
            SELECT leader_name, leader_phone, manager_name, manager_phone, 
                   member_name, member_phone, created_at, updated_at
            FROM duty_info 
            WHERE department = ? AND date = ?
        """, (department, date))
        row = c.fetchone()
        if not row:
            return {
                "leader_name": "",
                "leader_phone": "",
                "manager_name": "",
                "manager_phone": "",
                "member_name": "",
                "member_phone": "",
                "created_at": "",
                "updated_at": ""
            }
        return {
            "leader_name": row[0] or "",
            "leader_phone": row[1] or "",
            "manager_name": row[2] or "",
            "manager_phone": row[3] or "",
            "member_name": row[4] or "",
            "member_phone": row[5] or "",
            "created_at": row[6] or "",
            "updated_at": row[7] or ""
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"获取值班信息错误: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/duty-info/{department}")
async def update_department_duty_info(
    department: str,
    date: str,
    duty_info: DutyInfo,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """更新指定部门在指定日期的值班信息"""
    try:
        # 检查权限
        if not current_user.is_admin and current_user.department != department:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="没有权限更新该部门的值班信息"
            )
        
        c = db.cursor()
        # 检查记录是否存在
        c.execute("SELECT 1 FROM duty_info WHERE department = ? AND date = ?", (department, date))
        if c.fetchone():
            # 更新现有记录
            c.execute("""
                UPDATE duty_info 
                SET leader_name = ?, leader_phone = ?, 
                    manager_name = ?, manager_phone = ?,
                    member_name = ?, member_phone = ?,
                    updated_at = ?
                WHERE department = ? AND date = ?
            """, (
                duty_info.leader_name, duty_info.leader_phone,
                duty_info.manager_name, duty_info.manager_phone,
                duty_info.member_name, duty_info.member_phone,
                get_east8_time(),
                department, date
            ))
        else:
            # 插入新记录
            c.execute("""
                INSERT INTO duty_info 
                (department, date, leader_name, leader_phone, manager_name, 
                 manager_phone, member_name, member_phone, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                department, date,
                duty_info.leader_name, duty_info.leader_phone,
                duty_info.manager_name, duty_info.manager_phone,
                duty_info.member_name, duty_info.member_phone,
                get_east8_time(), get_east8_time()
            ))
        
        db.commit()
        return {"message": "值班信息更新成功"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新值班信息失败: {str(e)}"
        )

# 用户管理相关模型
class UserCreate(BaseModel):
    username: str
    password: str
    department: Optional[str] = None
    is_admin: bool = False

class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    department: Optional[str] = None
    is_admin: Optional[bool] = None

# 用户管理API
@app.get("/api/users", response_model=List[User])
async def get_users(
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限访问")
    
    try:
        cursor = db.cursor()
        cursor.execute("SELECT username, department, is_admin FROM users")
        users = [{"username": row[0], "department": row[1], "is_admin": bool(row[2])} for row in cursor.fetchall()]
        return users
    except sqlite3.Error as e:
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")

@app.get("/api/users/{username}", response_model=User)
async def get_user(username: str, current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限访问")
    conn = sqlite3.connect(DATABASE, check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT username, department, is_admin FROM users WHERE username = ?", (username,))
    user = cursor.fetchone()
    conn.close()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    return {"username": user[0], "department": user[1], "is_admin": bool(user[2])}

@app.post("/api/users")
async def create_user(
    user: UserCreate,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限创建用户")
    
    try:
        cursor = db.cursor()
        
        # 检查用户名是否已存在
        cursor.execute("SELECT username FROM users WHERE username = ?", (user.username,))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="用户名已存在")
        
        # 创建新用户
        hashed_password = bcrypt.hashpw(user.password.encode(), bcrypt.gensalt())
        current_time = get_east8_time()
        cursor.execute("""
            INSERT INTO users (username, password, department, is_admin, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user.username, hashed_password, user.department, user.is_admin, current_time, current_time))
        
        db.commit()
        return {"message": "用户创建成功"}
    except sqlite3.Error as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")

@app.put("/api/users/{username}")
async def update_user(username: str, user: UserUpdate, current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限修改用户")
    
    conn = None
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 检查用户是否存在
        cursor.execute("SELECT username FROM users WHERE username = ?", (username,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="用户不存在")
        
        # 检查新用户名是否已存在
        if user.username and user.username != username:
            cursor.execute("SELECT username FROM users WHERE username = ?", (user.username,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="用户名已存在")
        
        # 更新用户信息
        update_fields = []
        params = []
        
        if user.username is not None:
            update_fields.append("username = ?")
            params.append(user.username)
        
        if user.department is not None:
            update_fields.append("department = ?")
            params.append(user.department)
        
        if user.is_admin is not None:
            update_fields.append("is_admin = ?")
            params.append(user.is_admin)
            
        if user.password is not None:
            # 验证新密码长度
            if len(user.password) < 6:
                raise HTTPException(status_code=400, detail="密码长度不能少于6位")
            # 对密码进行加密
            hashed_password = bcrypt.hashpw(user.password.encode(), bcrypt.gensalt())
            update_fields.append("password = ?")
            params.append(hashed_password)
        
        if update_fields:
            update_fields.append("updated_at = ?")
            params.append(get_east8_time())
            params.append(username)
            cursor.execute(f"""
                UPDATE users 
                SET {', '.join(update_fields)}
                WHERE username = ?
            """, params)
            
            conn.commit()
        
        return {"message": "用户信息更新成功"}
    except sqlite3.Error as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")
    finally:
        if conn:
            conn.close()

@app.delete("/api/users/{username}")
async def delete_user(username: str, current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限删除用户")
    
    if username == "admin":
        raise HTTPException(status_code=400, detail="不能删除管理员账号")
    
    conn = None
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 检查用户是否存在
        cursor.execute("SELECT username FROM users WHERE username = ?", (username,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="用户不存在")
        
        # 删除用户
        cursor.execute("DELETE FROM users WHERE username = ?", (username,))
        conn.commit()
        return {"message": "用户删除成功"}
    except sqlite3.Error as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")
    finally:
        if conn:
            conn.close()

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

@app.post("/api/change-password")
async def change_password(
    request: ChangePasswordRequest,
    current_user: User = Depends(get_current_user)
):
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 获取当前用户的密码
        cursor.execute("SELECT password FROM users WHERE username = ?", (current_user.username,))
        user = cursor.fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")
        
        # 验证当前密码
        if not verify_password(request.current_password, user[0]):
            raise HTTPException(status_code=400, detail="当前密码错误")
        
        # 验证新密码长度
        if len(request.new_password) < 6:
            raise HTTPException(status_code=400, detail="新密码长度不能少于6位")
        
        # 更新密码
        hashed_password = bcrypt.hashpw(request.new_password.encode(), bcrypt.gensalt())
        current_time = get_east8_time()
        cursor.execute("""
            UPDATE users 
            SET password = ?, updated_at = ?
            WHERE username = ?
        """, (hashed_password, current_time, current_user.username))
        
        conn.commit()
        return {"message": "密码修改成功"}
    except sqlite3.Error as e:
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")
    finally:
        if conn:
            conn.close()

# 导入值班信息
@app.post("/api/import-duty-info")
async def import_duty_info(
    file: UploadFile = File(...),
    date: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限导入值班信息")
    
    try:
        # 读取 Excel 文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        cursor = db.cursor()
        
        # 开始事务
        cursor.execute("BEGIN TRANSACTION")
        
        # 遍历数据并插入
        for _, row in df.iterrows():
            department = row['部门']
            leader_name = row['值班领导姓名']
            leader_phone = row['值班领导电话']
            manager_name = row['值班经理姓名']
            manager_phone = row['值班经理电话']
            member_name = row['值班人员姓名']
            member_phone = row['值班人员电话']
            
            # 检查是否已存在该日期的值班信息
            cursor.execute("""
                SELECT 1 FROM duty_info 
                WHERE department = ? AND date = ?
            """, (department, date))
            
            if cursor.fetchone():
                # 更新现有记录
                cursor.execute("""
                    UPDATE duty_info 
                    SET leader_name = ?, leader_phone = ?,
                        manager_name = ?, manager_phone = ?,
                        member_name = ?, member_phone = ?,
                        updated_at = ?
                    WHERE department = ? AND date = ?
                """, (
                    leader_name, leader_phone,
                    manager_name, manager_phone,
                    member_name, member_phone,
                    get_east8_time(),
                    department, date
                ))
            else:
                # 插入新记录
                cursor.execute("""
                    INSERT INTO duty_info (
                        department, date, leader_name, leader_phone,
                        manager_name, manager_phone, member_name, member_phone,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    department, date, leader_name, leader_phone,
                    manager_name, manager_phone, member_name, member_phone,
                    get_east8_time(), get_east8_time()
                ))
        
        # 提交事务
        db.commit()
        return {"message": "导入成功"}
    except sqlite3.Error as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

# 导出值班信息
@app.get("/api/export-duty-info")
async def export_duty_info(
    date: str,
    current_user: User = Depends(get_current_user)
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限导出数据")
    
    try:
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        cursor = conn.cursor()
        
        # 获取指定日期的值班信息
        cursor.execute("""
            SELECT DISTINCT department, leader_name, leader_phone,
                   manager_name, manager_phone,
                   member_name, member_phone
            FROM duty_info
            WHERE date = ?
            ORDER BY department
        """, (date,))
        
        rows = cursor.fetchall()
        conn.close()
        
        # 创建DataFrame
        df = pd.DataFrame(rows, columns=[
            '部门', '值班领导姓名', '值班领导电话',
            '值班经理姓名', '值班经理电话',
            '值班人员姓名', '值班人员电话'
        ])
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            # 将DataFrame写入Excel文件
            df.to_excel(tmp.name, index=False, engine='openpyxl')
            
            # 返回文件
            return FileResponse(
                tmp.name,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=f'值班信息_{date}.xlsx'
            )
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

# 导出导入模板
@app.get("/api/export-template")
async def export_template(current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="没有权限导出模板")
    
    try:
        # 创建示例数据
        example_data = {
            '部门': ['示例部门1', '示例部门2'],
            '值班领导姓名': ['张三', '李四'],
            '值班领导电话': ['13800138000', '13900139000'],
            '值班经理姓名': ['王五', '赵六'],
            '值班经理电话': ['13700137000', '13600136000'],
            '值班人员姓名': ['钱七', '孙八'],
            '值班人员电话': ['13500135000', '13400134000']
        }
        
        # 创建DataFrame
        df = pd.DataFrame(example_data)
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            # 将DataFrame写入Excel文件
            df.to_excel(tmp.name, index=False, engine='openpyxl')
            
            # 返回文件
            return FileResponse(
                tmp.name,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename='值班信息导入模板.xlsx'
            )
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出模板失败: {str(e)}")

# 创建一个不需要认证的依赖
async def get_db_without_auth():
    conn = sqlite3.connect(DATABASE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

@app.get("/api/operation-center-duty")
async def get_operation_center_duty(
    date: str,
    db: sqlite3.Connection = Depends(get_db)
):
    """获取运营中心值班人员信息"""
    try:
        c = db.cursor()
        
        # 获取固定值班人员信息
        c.execute("""
            SELECT shift, leader_name, leader_phone, manager_name, manager_phone,
                   member_name, member_phone
            FROM operation_center_fixed_duty
        """)
        fixed_duty = {}
        for row in c.fetchall():
            fixed_duty[row[0]] = {
                "leader_name": row[1] or "",
                "leader_phone": row[2] or "",
                "manager_name": row[3] or "",
                "manager_phone": row[4] or "",
                "member_name": row[5] or "",
                "member_phone": row[6] or ""
            }
        
        # 获取备班人员信息
        c.execute("""
            SELECT shift, backup_name, backup_phone
            FROM operation_center_duty
            WHERE date = ?
        """, (date,))
        backup_duty = {}
        for row in c.fetchall():
            # 确保备班人员信息使用\n作为分隔符
            backup_name = row[1] or ""
            backup_phone = row[2] or ""
            if "、" in backup_name:
                backup_name = backup_name.replace("、", "\n")
            if "、" in backup_phone:
                backup_phone = backup_phone.replace("、", "\n")
            
            backup_duty[row[0]] = {
                "backup_name": backup_name,
                "backup_phone": backup_phone
            }
        
        # 如果没有找到当天的备班人员信息，尝试获取最近的有效备班人员信息
        if not backup_duty:
            c.execute("""
                SELECT shift, backup_name, backup_phone
                FROM operation_center_duty
                WHERE date < ?
                ORDER BY date DESC, shift
                LIMIT 4
            """, (date,))
            for row in c.fetchall():
                if row[0] not in backup_duty:
                    # 确保备班人员信息使用\n作为分隔符
                    backup_name = row[1] or ""
                    backup_phone = row[2] or ""
                    if "、" in backup_name:
                        backup_name = backup_name.replace("、", "\n")
                    if "、" in backup_phone:
                        backup_phone = backup_phone.replace("、", "\n")
                    
                    backup_duty[row[0]] = {
                        "backup_name": backup_name,
                        "backup_phone": backup_phone
                    }
        
        # 合并信息
        result = {}
        for shift in range(1, 5):
            result[shift] = {
                **fixed_duty.get(shift, {}),
                **backup_duty.get(shift, {})
            }
        
        return result
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取运营中心值班信息失败: {str(e)}"
        )

@app.post("/api/operation-center-duty")
async def create_operation_center_duty(
    duty_info: OperationCenterDuty,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """创建或更新运营中心值班人员信息"""
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="没有权限修改运营中心值班信息"
        )
    
    try:
        c = db.cursor()
        # 检查记录是否存在
        c.execute("""
            SELECT 1 FROM operation_center_duty 
            WHERE date = ? AND shift = ?
        """, (duty_info.date, duty_info.shift))
        
        if c.fetchone():
            # 更新现有记录
            c.execute("""
                UPDATE operation_center_duty 
                SET backup_name = ?, backup_phone = ?,
                    updated_at = ?
                WHERE date = ? AND shift = ?
            """, (
                duty_info.backup_name,
                duty_info.backup_phone,
                get_east8_time(),
                duty_info.date,
                duty_info.shift
            ))
        else:
            # 插入新记录
            c.execute("""
                INSERT INTO operation_center_duty 
                (date, shift, backup_name, backup_phone, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                duty_info.date,
                duty_info.shift,
                duty_info.backup_name,
                duty_info.backup_phone,
                get_east8_time(),
                get_east8_time()
            ))
        
        db.commit()
        return {"message": "运营中心值班信息更新成功"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新运营中心值班信息失败: {str(e)}"
        )

@app.post("/api/operation-center-fixed-duty")
async def create_operation_center_fixed_duty(
    duty_info: OperationCenterFixedDuty,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """创建或更新运营中心固定值班人员信息"""
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="没有权限修改运营中心固定值班信息"
        )
    
    try:
        c = db.cursor()
        # 检查记录是否存在
        c.execute("SELECT 1 FROM operation_center_fixed_duty WHERE shift = ?", (duty_info.shift,))
        
        if c.fetchone():
            # 更新现有记录
            c.execute("""
                UPDATE operation_center_fixed_duty 
                SET leader_name = ?, leader_phone = ?,
                    manager_name = ?, manager_phone = ?,
                    member_name = ?, member_phone = ?,
                    updated_at = ?
                WHERE shift = ?
            """, (
                duty_info.leader_name,
                duty_info.leader_phone,
                duty_info.manager_name,
                duty_info.manager_phone,
                duty_info.member_name,
                duty_info.member_phone,
                get_east8_time(),
                duty_info.shift
            ))
        else:
            # 插入新记录
            c.execute("""
                INSERT INTO operation_center_fixed_duty 
                (shift, leader_name, leader_phone, manager_name, 
                 manager_phone, member_name, member_phone,
                 created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                duty_info.shift,
                duty_info.leader_name,
                duty_info.leader_phone,
                duty_info.manager_name,
                duty_info.manager_phone,
                duty_info.member_name,
                duty_info.member_phone,
                get_east8_time(),
                get_east8_time()
            ))
        
        db.commit()
        return {"message": "运营中心固定值班信息更新成功"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新运营中心固定值班信息失败: {str(e)}"
        )

@app.get("/api/operation-center-duty/history")
async def get_operation_center_duty_history(
    start_date: str,
    end_date: str,
    shift: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """获取运营中心值班历史记录"""
    try:
        c = db.cursor()
        query = """
            SELECT date, shift, backup_name, backup_phone, created_at, updated_at
            FROM operation_center_duty
            WHERE date BETWEEN ? AND ?
        """
        params = [start_date, end_date]
        
        if shift:
            query += " AND shift = ?"
            params.append(shift)
        
        query += " ORDER BY date DESC, shift"
        
        c.execute(query, params)
        history = []
        for row in c.fetchall():
            # 获取对应班次的固定值班人员信息
            c.execute("""
                SELECT leader_name, leader_phone, manager_name, manager_phone,
                       member_name, member_phone
                FROM operation_center_fixed_duty
                WHERE shift = ?
            """, (row[1],))
            fixed_duty = c.fetchone()
            
            if fixed_duty:
                history.append({
                    "date": row[0],
                    "shift": row[1],
                    "leader_name": fixed_duty[0],
                    "leader_phone": fixed_duty[1],
                    "manager_name": fixed_duty[2],
                    "manager_phone": fixed_duty[3],
                    "member_name": fixed_duty[4],
                    "member_phone": fixed_duty[5],
                    "backup_name": row[2],
                    "backup_phone": row[3],
                    "created_at": row[4],
                    "updated_at": row[5]
                })
        
        return history
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取运营中心值班历史记录失败: {str(e)}"
        )

@app.post("/api/operation-center-duty/batch")
async def batch_update_operation_center_duty(
    duty_info: OperationCenterDutyBatch,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """批量更新运营中心固定值班人员信息"""
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="没有权限修改运营中心固定值班信息"
        )
    
    try:
        c = db.cursor()
        current_time = get_east8_time()
        
        for shift in duty_info.shifts:
            c.execute("""
                SELECT 1 FROM operation_center_fixed_duty WHERE shift = ?
            """, (shift,))
            
            if c.fetchone():
                c.execute("""
                    UPDATE operation_center_fixed_duty 
                    SET leader_name = ?, leader_phone = ?,
                        manager_name = ?, manager_phone = ?,
                        member_name = ?, member_phone = ?,
                        updated_at = ?
                    WHERE shift = ?
                """, (
                    duty_info.leader_name,
                    duty_info.leader_phone,
                    duty_info.manager_name,
                    duty_info.manager_phone,
                    duty_info.member_name,
                    duty_info.member_phone,
                    current_time,
                    shift
                ))
            else:
                c.execute("""
                    INSERT INTO operation_center_fixed_duty 
                    (shift, leader_name, leader_phone, manager_name, 
                     manager_phone, member_name, member_phone,
                     created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    shift,
                    duty_info.leader_name,
                    duty_info.leader_phone,
                    duty_info.manager_name,
                    duty_info.manager_phone,
                    duty_info.member_name,
                    duty_info.member_phone,
                    current_time,
                    current_time
                ))
        
        db.commit()
        return {"message": "运营中心固定值班信息批量更新成功"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新运营中心固定值班信息失败: {str(e)}"
        )

@app.get("/api/operation-center-duty/export")
async def export_operation_center_duty(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """导出运营中心值班信息"""
    try:
        c = db.cursor()
        # 获取固定值班人员信息
        c.execute("""
            SELECT shift, leader_name, leader_phone, manager_name, manager_phone,
                   member_name, member_phone
            FROM operation_center_fixed_duty
            ORDER BY shift
        """)
        fixed_duty = {}
        for row in c.fetchall():
            fixed_duty[row[0]] = {
                "leader_name": row[1],
                "leader_phone": row[2],
                "manager_name": row[3],
                "manager_phone": row[4],
                "member_name": row[5],
                "member_phone": row[6]
            }
        
        # 获取备班人员信息
        c.execute("""
            SELECT date, shift, backup_name, backup_phone
            FROM operation_center_duty
            WHERE date BETWEEN ? AND ?
            ORDER BY date, shift
        """, (start_date, end_date))
        
        # 创建 Excel 文件
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()
        
        # 设置表头
        headers = ['日期', '班次', '值班领导', '领导电话', '值班主管', '主管电话',
                  '值班人员', '人员电话', '备班人员', '备班电话']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        
        # 写入数据
        row = 1
        for record in c.fetchall():
            date = record[0]
            shift = record[1]
            fixed_info = fixed_duty.get(shift, {})
            
            worksheet.write(row, 0, date)
            worksheet.write(row, 1, f"{shift}班")
            worksheet.write(row, 2, fixed_info.get('leader_name', ''))
            worksheet.write(row, 3, fixed_info.get('leader_phone', ''))
            worksheet.write(row, 4, fixed_info.get('manager_name', ''))
            worksheet.write(row, 5, fixed_info.get('manager_phone', ''))
            worksheet.write(row, 6, fixed_info.get('member_name', ''))
            worksheet.write(row, 7, fixed_info.get('member_phone', ''))
            worksheet.write(row, 8, record[2] or '')
            worksheet.write(row, 9, record[3] or '')
            row += 1
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=operation_center_duty_{start_date}_{end_date}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导出运营中心值班信息失败: {str(e)}"
        )

@app.post("/api/operation-center-duty/import")
async def import_operation_center_duty(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: sqlite3.Connection = Depends(get_db)
):
    """导入运营中心值班信息"""
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="没有权限导入运营中心值班信息"
        )
    
    try:
        # 读取 Excel 文件
        contents = await file.read()
        try:
            df = pd.read_excel(io.BytesIO(contents))
        except Exception as e:
            print(f"读取Excel文件错误: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"读取Excel文件失败: {str(e)}"
            )
        
        # 检查必要的列是否存在
        required_columns = ['日期', '班次', '值班领导', '领导电话', '值班主管', '主管电话', '值班人员', '人员电话', '备班人员', '备班电话']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}"
            )
        
        c = db.cursor()
        current_time = get_east8_time()
        
        # 处理每一行数据
        for index, row in df.iterrows():
            try:
                # 检查日期格式
                date = str(row['日期'])
                try:
                    # 尝试解析日期，允许带时间的格式
                    parsed_date = pd.to_datetime(date)
                    date = parsed_date.date().isoformat()  # 转换为 YYYY-MM-DD 格式
                except ValueError:
                    # 如果解析失败，检查是否符合 YYYY-MM-DD 格式
                    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
                        raise ValueError(f"第{index+1}行日期格式错误: {date}")
                
                # 检查班次格式
                shift_str = str(row['班次'])
                if not re.match(r'^\d+班$', shift_str):
                    raise ValueError(f"第{index+1}行班次格式错误: {shift_str}")
                shift = int(shift_str.replace('班', ''))
                
                # 更新或插入固定值班人员信息
                c.execute("""
                    SELECT 1 FROM operation_center_fixed_duty WHERE shift = ?
                """, (shift,))
                
                # 处理姓名和电话号码，将顿号分隔转换为换行分隔
                def process_names_and_phones(names_str, phones_str):
                    if not names_str or not phones_str:
                        return "", ""
                    names = [name.strip() for name in names_str.split('、')]
                    phones = [phone.strip() for phone in phones_str.split('、')]
                    if len(names) != len(phones):
                        raise ValueError(f"姓名和电话号码数量不匹配: {names_str} vs {phones_str}")
                    return '\n'.join(names), '\n'.join(phones)
                
                leader_names, leader_phones = process_names_and_phones(str(row['值班领导']), str(row['领导电话']))
                manager_names, manager_phones = process_names_and_phones(str(row['值班主管']), str(row['主管电话']))
                member_names, member_phones = process_names_and_phones(str(row['值班人员']), str(row['人员电话']))
                backup_names, backup_phones = process_names_and_phones(str(row['备班人员']), str(row['备班电话']))
                
                if c.fetchone():
                    c.execute("""
                        UPDATE operation_center_fixed_duty 
                        SET leader_name = ?, leader_phone = ?,
                            manager_name = ?, manager_phone = ?,
                            member_name = ?, member_phone = ?,
                            updated_at = ?
                        WHERE shift = ?
                    """, (
                        leader_names,
                        leader_phones,
                        manager_names,
                        manager_phones,
                        member_names,
                        member_phones,
                        current_time,
                        shift
                    ))
                else:
                    c.execute("""
                        INSERT INTO operation_center_fixed_duty 
                        (shift, leader_name, leader_phone, manager_name, 
                         manager_phone, member_name, member_phone,
                         created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        shift,
                        leader_names,
                        leader_phones,
                        manager_names,
                        manager_phones,
                        member_names,
                        member_phones,
                        current_time,
                        current_time
                    ))
                
                # 更新或插入备班人员信息
                c.execute("""
                    SELECT 1 FROM operation_center_duty 
                    WHERE date = ? AND shift = ?
                """, (date, shift))
                
                if c.fetchone():
                    c.execute("""
                        UPDATE operation_center_duty 
                        SET backup_name = ?, backup_phone = ?,
                            updated_at = ?
                        WHERE date = ? AND shift = ?
                    """, (
                        backup_names,
                        backup_phones,
                        current_time,
                        date,
                        shift
                    ))
                else:
                    c.execute("""
                        INSERT INTO operation_center_duty 
                        (date, shift, backup_name, backup_phone, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        date,
                        shift,
                        backup_names,
                        backup_phones,
                        current_time,
                        current_time
                    ))
            except Exception as e:
                print(f"处理第{index+1}行数据时出错: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"处理第{index+1}行数据时出错: {str(e)}"
                )
        
        db.commit()
        return {"message": "运营中心值班信息导入成功"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"导入运营中心值班信息错误: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入运营中心值班信息失败: {str(e)}"
        )

# 登出接口
@app.post("/api/logout")
async def logout():
    return {"message": "登出成功"}

@app.get("/api/operation_center/template")
async def download_operation_center_template():
    """下载运营中心值班人员信息导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "运营中心值班人员"
    
    # 设置表头
    headers = ['日期', '班次', '值班领导', '领导电话', '值班主管', '主管电话', 
              '值班人员', '人员电话', '备班人员', '备班电话']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 设置示例数据（每个班次3人）
    example_data = [
        ['2024-04-01', '一班', '张三、李四、王五', '13800000001、13800000002、13800000003', 
         '赵六、钱七、孙八', '13800000004、13800000005、13800000006',
         '周九、吴十、郑十一', '13800000007、13800000008、13800000009',
         '备班1、备班2、备班3', '13900000001、13900000002、13900000003'],
        ['2024-04-02', '二班', '张十二、李十三、王十四', '13800000011、13800000012、13800000013',
         '赵十五、钱十六、孙十七', '13800000014、13800000015、13800000016',
         '周十八、吴十九、郑二十', '13800000017、13800000018、13800000019',
         '备班4、备班5、备班6', '13900000004、13900000005、13900000006']
    ]
    
    # 写入示例数据
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 添加说明行
    ws.cell(row=4, column=1, value="说明：")
    ws.cell(row=5, column=1, value="1. 每个班次可以设置多名值班人员，用顿号（、）分隔")
    ws.cell(row=6, column=1, value="2. 姓名和电话必须一一对应，且数量相同")
    ws.cell(row=7, column=1, value="3. 备班人员信息为选填项")
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # 保存到内存中
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = quote("运营中心值班人员导入模板.xlsx")
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 